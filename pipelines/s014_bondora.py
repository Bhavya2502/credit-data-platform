"""S-014 · Bondora/Go&Grow — fetch → bronze → silver → reconcile.

789k European consumer loans carrying BOTH a Basel-shaped 12-month default
target and realised repayment amounts, so PD and LGD can be modelled from one
table. Complements S-079: corporate workout recoveries there, retail here.

Usage
-----
    python pipelines/s014_bondora.py            # fetch and load
    python pipelines/s014_bondora.py --check    # HEAD only, report freshness
"""

from __future__ import annotations

import argparse
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.common import catalog  # noqa: E402
from src.common.settings import Settings  # noqa: E402
from src.common.storage import BronzeStore  # noqa: E402
from src.connectors.bulk_file import cleanup, download, probe  # noqa: E402

SOURCE_ID = "S-014"
CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "sources" / "S-014_bondora.yaml"
SILVER_TABLE = "silver.bondora_loans"
DICT_TABLE = "silver.bondora_dictionary"

BOOL_COLS = ["is_default", "is_early_repaid_within_14_days", "has_default_within_12_months"]
DATE_COLS = ["loan_issued_at", "early_repaid_at", "loan_last_recorded_action_date_local",
             "next_payment_date_local", "debt_occured_date_local"]
NUM_COLS = ["issued_amount", "initial_interest_rate", "nr_of_payments", "principal_balance",
            "principal_debt", "principal_paid_total", "interest_paid_total",
            "extra_interest_paid_total", "late_fee_paid_total", "maintenance_fee_paid_total",
            "next_payment_nr", "days_past_due_principal", "months_in_default", "months_on_book",
            "repaid_amount_total", "initial_loan_duration", "combined_income",
            "projected_npv_return"]


def load_config() -> dict:
    with CONFIG_PATH.open(encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def read_workbook(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Read the loan sheet and the file's own data dictionary.

    Uses openpyxl in read-only mode and streams rows; loading a 155 MB workbook
    through pandas' default path is far heavier.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        data = [r for r in rows if r is not None and any(v is not None for v in r)]
        loans = pd.DataFrame(data, columns=header)

        dictionary = pd.DataFrame(columns=["column", "description"])
        if "Dataset Dictionary" in wb.sheetnames:
            dws = wb["Dataset Dictionary"]
            drows = [
                (str(r[0]).strip(), str(r[1]).strip())
                for r in dws.iter_rows(values_only=True)
                if r and r[0] and r[1] and str(r[0]).strip().lower() != "column"
            ]
            dictionary = pd.DataFrame(drows, columns=["column", "description"])
    finally:
        wb.close()
    return loans, dictionary


def normalise(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Type the columns, preserving NULL semantics that matter for modelling."""
    notes: list[str] = []

    for col in NUM_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    for col in DATE_COLS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    for col in BOOL_COLS:
        if col in df.columns:
            # NULL must survive. For has_default_within_12_months NULL means
            # "not yet measurable", NOT "no default" — coercing it to False
            # understates the default rate and biases any PD model.
            mapped = df[col].astype(str).str.strip().str.lower().map(
                {"true": True, "false": False, "1": True, "0": False}
            )
            df[col] = mapped.astype("boolean")

    if "has_default_within_12_months" in df.columns:
        n_null = int(df["has_default_within_12_months"].isna().sum())
        notes.append(f"{n_null:,} loans not yet measurable for 12-month default (kept NULL)")

    df["_source_id"] = SOURCE_ID
    df["_fetched_at"] = datetime.now(timezone.utc)
    return df, notes


def reconcile(con, cfg: dict) -> list[tuple[str, bool, str]]:
    results: list[tuple[str, bool, str]] = []

    total = con.execute(f"SELECT count(*) FROM {SILVER_TABLE}").fetchone()[0]
    expected = int(cfg["coverage"]["rows_verified_2026_08"])
    tol = float([c for c in cfg["reconciliation"] if c["check"] == "row_count"][0]["tolerance"])
    within = abs(total - expected) / expected <= tol
    results.append((
        "row_count", within,
        f"{total:,} rows vs {expected:,} expected (+/-{tol:.0%}); file grows as loans are issued",
    ))

    row = con.execute(
        f"""
        SELECT count(*) FILTER (WHERE has_default_within_12_months IS NOT NULL),
               avg(CASE WHEN has_default_within_12_months THEN 1.0 ELSE 0.0 END)
                 FILTER (WHERE has_default_within_12_months IS NOT NULL)
        FROM {SILVER_TABLE}
        """
    ).fetchone()
    measurable, rate = row[0] or 0, row[1]
    if rate is not None:
        ok = 0.02 <= rate <= 0.40
        results.append((
            "default_rate_plausible", ok,
            f"12-month default rate {rate:.2%} across {measurable:,} measurable loans",
        ))

    # Truth check: the lender's own grade should order realised defaults.
    grades = con.execute(
        f"""
        SELECT customer_risk_rating,
               count(*) AS n,
               avg(CASE WHEN has_default_within_12_months THEN 1.0 ELSE 0.0 END) AS dr
        FROM {SILVER_TABLE}
        WHERE has_default_within_12_months IS NOT NULL
          AND customer_risk_rating IS NOT NULL AND trim(customer_risk_rating) <> ''
        GROUP BY 1 HAVING count(*) >= 100 ORDER BY 1
        """
    ).fetchall()
    if len(grades) >= 3:
        labels = [g[0] for g in grades]
        rates = [g[2] for g in grades]
        # Spearman-style: does default rate rise as the grade worsens (A→HR)?
        ranked = sorted(range(len(rates)), key=lambda i: rates[i])
        concordant = sum(
            1 for a in range(len(rates)) for b in range(a + 1, len(rates))
            if rates[b] > rates[a]
        )
        pairs = len(rates) * (len(rates) - 1) / 2
        corr = (2 * concordant / pairs) - 1 if pairs else 0
        detail = " · ".join(f"{labels[i]}:{rates[i]:.1%}" for i in range(len(labels)))
        results.append((
            "risk_rating_orders_defaults", corr >= 0.7,
            f"rank concordance {corr:.2f} across {len(grades)} grades — {detail}",
        ))

    bad = con.execute(
        f"SELECT count(*) FROM {SILVER_TABLE} WHERE issued_amount IS NULL OR issued_amount <= 0"
    ).fetchone()[0]
    results.append((
        "issued_amount_positive", bad < max(1, total * 0.001),
        f"{bad:,} rows with non-positive issued amount",
    ))

    cov = con.execute(
        f"SELECT min(loan_issued_at), max(loan_issued_at), count(DISTINCT country) FROM {SILVER_TABLE}"
    ).fetchone()
    results.append((
        "coverage", True,
        f"{cov[0]} to {cov[1]} across {cov[2]} countries",
    ))
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Bondora pipeline (S-014)")
    parser.add_argument("--check", action="store_true", help="probe only, do not download")
    args = parser.parse_args()

    cfg = load_config()
    settings = Settings.from_env()
    run_id = uuid.uuid4().hex[:12]
    started = datetime.now(timezone.utc)

    print("=" * 70)
    print(f"S-014 · Bondora / Go&Grow   run {run_id}")
    print("=" * 70)

    url = cfg["download_url"]
    headers = cfg["fetch"]["headers"]

    info = probe(url, headers=headers)
    size_mb = (info.content_length or 0) / 1_048_576
    print(f"\n  remote    HTTP {info.status_code} · {size_mb:.1f} MB · modified {info.last_modified}")
    if args.check:
        return 0
    if info.status_code >= 400:
        print(f"  ✗ source unreachable (HTTP {info.status_code})")
        return 1

    dl = download(url, headers=headers, timeout=cfg["fetch"]["timeout_seconds"],
                  max_retries=cfg["fetch"]["retries"], filename="bondora_loan_dataset.xlsx")
    print(f"  fetched   {dl.size_bytes/1_048_576:.1f} MB in {dl.seconds}s · sha256 {dl.sha256[:12]}…")

    store = BronzeStore(settings)
    bronze = store.put_bytes(
        source_id=SOURCE_ID,
        filename="bondora_loan_dataset.xlsx",
        payload=dl.read_bytes(),
        source_url=url,
        content_type=dl.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        notes=f"last_modified={dl.last_modified}",
    )
    print(f"  bronze    {bronze.key}")

    loans, dictionary = read_workbook(dl.path)
    print(f"  parsed    {len(loans):,} loans x {len(loans.columns)} cols · "
          f"{len(dictionary)} dictionary entries")

    loans, notes = normalise(loans)
    for n in notes:
        print(f"  note      {n}")

    with catalog.connect(settings) as con:
        catalog.ensure_schemas(con)
        catalog.register_source(con, cfg)
        con.execute("CREATE SCHEMA IF NOT EXISTS silver")

        con.register("incoming_loans", loans)
        con.execute(f"CREATE OR REPLACE TABLE {SILVER_TABLE} AS SELECT * FROM incoming_loans")
        con.unregister("incoming_loans")
        print(f"  silver    {len(loans):,} rows → {SILVER_TABLE}")

        if len(dictionary):
            con.register("incoming_dict", dictionary)
            con.execute(f"CREATE OR REPLACE TABLE {DICT_TABLE} AS SELECT * FROM incoming_dict")
            con.unregister("incoming_dict")
            print(f"  silver    {len(dictionary)} rows → {DICT_TABLE} (field definitions)")

        catalog.log_run(
            con, run_id=run_id, source_id=SOURCE_ID, started_at=started, status="success",
            bronze_key=bronze.key, source_url=url,
            rows_fetched=len(loans), rows_loaded=len(loans),
            target_table=SILVER_TABLE, message=f"last_modified={dl.last_modified}",
        )

        print("\nReconciliation")
        print("-" * 70)
        results = reconcile(con, cfg)
        for name, passed, detail in results:
            print(f"  [{'PASS' if passed else 'FAIL'}] {name}: {detail}")
            catalog.record_qa(
                con, run_id=run_id, source_id=SOURCE_ID, check_name=name,
                observed=1.0 if passed else 0.0, expected=1.0, tolerance=None,
                passed=passed, detail=detail,
            )

        failed = [r for r in results if not r[1]]
        if failed:
            print(f"\n  {len(failed)} check(s) failed.\n")
            cleanup(dl)
            return 1

        con.execute(
            "UPDATE catalog.source_registry SET last_success_at = ? WHERE source_id = ?",
            [datetime.now(timezone.utc), SOURCE_ID],
        )
        print("\n  All checks passed.\n")

    cleanup(dl)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
